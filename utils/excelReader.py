import openpyxl
import pytest
import os
from pathlib import Path

class ExcelReader:

    @staticmethod
    def getRowCount(path,sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        return sheet.max_row

    @staticmethod
    def getColumnCount(path, sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        return sheet.max_column

    @staticmethod
    def getCellData(path, sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        max_rows= ExcelReader.getRowCount(path,sheetName)
        max_colmns = ExcelReader.getColumnCount(path,sheetName)
        mainList=[]

        for i in range(2, max_rows+1):
            dataList = []
            for j in range(1,max_colmns+1):
                data = sheet.cell(row=i,column=j).value
                dataList.insert(j,data) # Inserting all column values to a list
            mainList.insert(i, dataList)  # Inserting all row values to a list
        return mainList

    @staticmethod
    def setCellData(path, sheetName,rowNum,colNum,data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        sheet.cell(row=rowNum,column=colNum).value=data
        workbook.save(path)




